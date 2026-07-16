"""
文件名: tests/services/test_export_service.py
创建时间: 2026-06-26
作者: TalentSense Team
功能描述: ExportService 单元测试
"""
import io
import pytest
from unittest.mock import AsyncMock, MagicMock
from openpyxl import load_workbook
from app.services.export_service import ExportService


@pytest.fixture
def svc():
    s = ExportService()
    # motor find() 同步返回 cursor，用 MagicMock
    s.resumes_coll = MagicMock()
    s.resumes_coll.find.return_value.to_list = AsyncMock(return_value=[])
    return s


@pytest.mark.asyncio
async def test_export_excel_basic(svc):
    """AC14.1: 导出 Excel，含候选人列表"""
    svc.resumes_coll.find.return_value.to_list = AsyncMock(return_value=[
        {
            "resume_id": "r1", "candidate_id": "c1", "name": "张三", "gender": "男",
            "age": 28, "education": "本科", "work_years": 5,
            "skills": ["Java", "Spring"], "expected_salary": {"min": 20, "max": 30},
            "tags": ["已面试"], "phone_masked": "138****1234", "email_masked": "z***@example.com",
            "location": "北京"
        }
    ])
    excel_bytes = await svc.export_excel(resume_ids=["r1"], columns=["name", "skills", "work_years"])
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(1, 1).value == "姓名"
    assert ws.cell(2, 1).value == "张三"


@pytest.mark.asyncio
async def test_export_excel_empty(svc):
    """AC14.2: 空列表返回仅含表头"""
    svc.resumes_coll.find.return_value.to_list = AsyncMock(return_value=[])
    excel_bytes = await svc.export_excel(resume_ids=[], columns=["name"])
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(1, 1).value == "姓名"
    assert ws.max_row == 1


@pytest.mark.asyncio
async def test_export_excel_all_columns(svc):
    """AC14.3: 包含所有字段"""
    svc.resumes_coll.find.return_value.to_list = AsyncMock(return_value=[
        {"resume_id": "r1", "candidate_id": "c1", "name": "李四", "work_years": 3, "skills": ["Python"],
         "expected_salary": {"min": 15, "max": 25}, "phone_masked": "139****5678", "email_masked": "l***@x.com"}
    ])
    columns = ["name", "gender", "age", "education", "work_years", "skills",
               "expected_salary", "tags", "phone_masked", "email_masked", "location"]
    excel_bytes = await svc.export_excel(resume_ids=["r1"], columns=columns)
    wb = load_workbook(io.BytesIO(excel_bytes))
    assert wb.active.max_column == len(columns)


@pytest.mark.asyncio
async def test_export_by_resume_ids(svc):
    """导出应按 resume_id 查询（前端传的是 resume_id 而非 candidate_id）"""
    svc.resumes_coll.find.return_value.to_list = AsyncMock(return_value=[
        {
            "resume_id": "res_001", "candidate_id": "c1",
            "basic_info": {"name": "张三", "gender": "男", "age": 28},
            "education": "本科", "work_years": 5,
            "skills": ["Java"], "expected_salary": {"min": 20, "max": 30},
            "tags": ["已面试"], "location": "北京"
        }
    ])
    excel_bytes = await svc.export_excel(
        resume_ids=["res_001"],
        columns=["name", "work_years"]
    )
    # 验证查询用的是 resume_id 而非 candidate_id
    call_args = svc.resumes_coll.find.call_args
    query = call_args.args[0]
    assert "resume_id" in query
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(2, 1).value == "张三"
